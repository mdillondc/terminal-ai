import os
import re
import hashlib
from typing import List, Dict, Any, Optional, Tuple
import tiktoken
import pypdf
from settings_manager import SettingsManager
from rag_config import is_supported_file, get_file_type_info
import docx2txt
import xlrd
import openpyxl
from striprtf.striprtf import rtf_to_text
from odf.opendocument import load as odf_load
from odf.text import P
from odf.table import Table, TableRow, TableCell
from odf import text, teletype
from print_helper import print_md


class DocumentProcessor:
    def __init__(self):
        self.settings_manager = SettingsManager.getInstance()
        self.encoding = tiktoken.get_encoding("cl100k_base")  # Standard encoding

    def count_tokens(self, text: str) -> int:
        """Count tokens in text using tiktoken"""
        return len(self.encoding.encode(text))

    def load_file(self, file_path: str) -> str:
        """Load content from supported file types"""
        file_info = get_file_type_info(file_path)
        processor = file_info.get('processor', 'text')

        if processor == 'pdf':
            return self.load_pdf_file(file_path)
        elif processor == 'docx':
            return self.load_docx_file(file_path)
        elif processor == 'doc':
            return self.load_docx_file(file_path)  # docx2txt handles both formats
        elif processor == 'xlsx':
            return self.load_xlsx_file(file_path)
        elif processor == 'xls':
            return self.load_xls_file(file_path)
        elif processor == 'rtf':
            return self.load_rtf_file(file_path)
        elif processor == 'odt':
            return self.load_odt_file(file_path)
        elif processor == 'ods':
            return self.load_ods_file(file_path)
        elif processor == 'odp':
            return self.load_odp_file(file_path)
        elif processor == 'email':
            return self.load_email_file(file_path)
        else:
            return self.load_text_file(file_path)

    def load_text_file(self, file_path: str) -> str:
        """Load content from a text or markdown file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            return content
        except UnicodeDecodeError:
            # Try with different encoding if UTF-8 fails
            try:
                with open(file_path, 'r', encoding='latin-1') as f:
                    content = f.read()
                return content
            except Exception as e:
                print_md(f"Error reading file {file_path}: {e}")
                return ""
        except Exception as e:
            print_md(f"Error loading file {file_path}: {e}")
            return ""

    def load_pdf_file(self, file_path: str) -> str:
        """Load content from a PDF file"""
        try:
            content = []
            with open(file_path, 'rb') as file:
                pdf_reader = pypdf.PdfReader(file)

                for page_num, page in enumerate(pdf_reader.pages):
                    try:
                        page_text = page.extract_text()
                        if page_text.strip():
                            content.append(f"--- Page {page_num + 1} ---\n{page_text}")
                    except Exception as e:
                        print_md(f"Error extracting text from page {page_num + 1} of {file_path}: {e}")
                        continue

            return "\n\n".join(content)
        except Exception as e:
            print_md(f"Error loading PDF file {file_path}: {e}")
            return ""

    def load_docx_file(self, file_path: str) -> str:
        """Load content from a Word (.docx or .doc) file using docx2txt"""
        try:
            # docx2txt handles both .doc and .docx formats
            content = docx2txt.process(file_path)
            return content if content else ""
        except Exception as e:
            print_md(f"Error loading Word file {file_path}: {e}")
            return ""

    def load_xlsx_file(self, file_path: str) -> str:
        """Load content from an Excel (.xlsx) file"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            content = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content.append(f"--- Sheet: {sheet_name} ---")

                # Extract all non-empty cells
                sheet_content = []
                for row in sheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None and str(cell).strip():
                            row_text.append(str(cell).strip())
                    if row_text:
                        sheet_content.append(" | ".join(row_text))

                if sheet_content:
                    content.extend(sheet_content)
                else:
                    content.append("(empty sheet)")

            return "\n\n".join(content)
        except Exception as e:
            print_md(f"Error loading Excel file {file_path}: {e}")
            return ""

    def load_xls_file(self, file_path: str) -> str:
        """Load content from a legacy Excel (.xls) file"""
        try:
            workbook = xlrd.open_workbook(file_path)
            content = []

            for sheet_idx in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_idx)
                sheet_name = sheet.name
                content.append(f"--- Sheet: {sheet_name} ---")

                # Extract all non-empty cells
                sheet_content = []
                for row_idx in range(sheet.nrows):
                    row_text = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell(row_idx, col_idx).value
                        if cell_value is not None and str(cell_value).strip():
                            row_text.append(str(cell_value).strip())
                    if row_text:
                        sheet_content.append(" | ".join(row_text))

                if sheet_content:
                    content.extend(sheet_content)
                else:
                    content.append("(empty sheet)")

            return "\n\n".join(content)
        except Exception as e:
            print_md(f"Error loading legacy Excel file {file_path}: {e}")
            return ""

    def load_rtf_file(self, file_path: str) -> str:
        """Load content from an RTF file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                rtf_content = file.read()

            # Convert RTF to plain text
            plain_text = rtf_to_text(rtf_content)
            return plain_text
        except UnicodeDecodeError:
            # Try with different encoding if UTF-8 fails
            try:
                with open(file_path, 'r', encoding='latin-1') as file:
                    rtf_content = file.read()
                plain_text = rtf_to_text(rtf_content)
                return plain_text
            except Exception as e:
                print_md(f"Error reading RTF file {file_path}: {e}")
                return ""
        except Exception as e:
            print_md(f"Error loading RTF file {file_path}: {e}")
            return ""

    def load_odt_file(self, file_path: str) -> str:
        """Load content from an OpenDocument Text (.odt) file"""
        try:
            doc = odf_load(file_path)
            content = []

            # Extract all paragraphs
            for paragraph in doc.getElementsByType(P):
                para_text = teletype.extractText(paragraph)
                if para_text.strip():
                    content.append(para_text.strip())

            # Extract table content
            for table in doc.getElementsByType(Table):
                content.append("--- Table ---")
                for row in table.getElementsByType(TableRow):
                    row_text = []
                    for cell in row.getElementsByType(TableCell):
                        cell_text = teletype.extractText(cell)
                        if cell_text.strip():
                            row_text.append(cell_text.strip())
                    if row_text:
                        content.append(" | ".join(row_text))

            return "\n\n".join(content)
        except Exception as e:
            print_md(f"Error loading ODT file {file_path}: {e}")
            return ""

    def load_ods_file(self, file_path: str) -> str:
        """Load content from an OpenDocument Spreadsheet (.ods) file"""
        try:
            from odf.table import Table, TableRow, TableCell
            from odf.text import P

            doc = odf_load(file_path)
            content = []

            # Extract all sheets (tables in ODS)
            for table in doc.getElementsByType(Table):
                table_name = table.getAttribute('name') or 'Sheet'
                content.append(f"--- Sheet: {table_name} ---")

                sheet_content = []
                for row in table.getElementsByType(TableRow):
                    row_text = []
                    for cell in row.getElementsByType(TableCell):
                        cell_text = teletype.extractText(cell)
                        if cell_text.strip():
                            row_text.append(cell_text.strip())
                    if row_text:
                        sheet_content.append(" | ".join(row_text))

                if sheet_content:
                    content.extend(sheet_content)
                else:
                    content.append("(empty sheet)")

            return "\n\n".join(content)
        except Exception as e:
            print_md(f"Error loading ODS file {file_path}: {e}")
            return ""

    def load_odp_file(self, file_path: str) -> str:
        """Load content from an OpenDocument Presentation (.odp) file"""
        try:
            from odf.draw import Page
            from odf.text import P

            doc = odf_load(file_path)
            content = []

            # Extract all slides (pages in ODP)
            pages = doc.getElementsByType(Page)
            for i, page in enumerate(pages, 1):
                content.append(f"--- Slide {i} ---")

                # Extract all text from the slide
                slide_content = []
                for paragraph in page.getElementsByType(P):
                    para_text = teletype.extractText(paragraph)
                    if para_text.strip():
                        slide_content.append(para_text.strip())

                if slide_content:
                    content.extend(slide_content)
                else:
                    content.append("(no text content)")

            return "\n\n".join(content)
        except Exception as e:
            print_md(f"Error loading ODP file {file_path}: {e}")
            return ""

    def load_email_file(self, file_path: str) -> str:
        """Load content from an email file, skipping attachments"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
        except UnicodeDecodeError:
            try:
                with open(file_path, 'r', encoding='latin-1') as f:
                    content = f.read()
            except Exception as e:
                print_md(f"Error reading email file {file_path}: {e}")
                return ""
        except Exception as e:
            print_md(f"Error loading email file {file_path}: {e}")
            return ""

        # Parse email content and extract only text parts
        return self._parse_email_content(content)

    def _parse_email_content(self, content: str) -> str:
        """Parse email content and extract only text parts, skipping attachments"""
        lines = content.split('\n')
        result_lines = []

        # Extract headers until we find the first empty line
        header_section = True
        boundary = None
        current_section_type = None
        skip_current_section = False

        for line in lines:
            if header_section:
                if line.strip() == '':
                    header_section = False
                    result_lines.append(line)
                    continue

                # Extract boundary from Content-Type header
                if line.startswith('Content-Type:') and 'boundary=' in line:
                    boundary_start = line.find('boundary=') + 9
                    boundary_end = line.find(';', boundary_start)
                    if boundary_end == -1:
                        boundary_end = len(line)
                    boundary = line[boundary_start:boundary_end].strip().strip('"')

                result_lines.append(line)
                continue

            # If we have a boundary, check for MIME parts
            if boundary and line.startswith('--' + boundary):
                # Reset section state
                current_section_type = None
                skip_current_section = False
                result_lines.append(line)
                continue

            # Check content headers within MIME parts
            if line.startswith('Content-Type:'):
                current_section_type = line.lower()
                result_lines.append(line)
                continue

            if line.startswith('Content-Transfer-Encoding:'):
                if 'base64' in line.lower():
                    skip_current_section = True
                result_lines.append(line)
                continue

            # Skip base64 encoded content (attachments)
            if skip_current_section and line.strip() and not line.startswith('Content-') and not line.startswith('--'):
                # This is likely base64 content, skip it
                continue

            # Include all other content
            result_lines.append(line)

        return '\n'.join(result_lines)

    def clean_text(self, text: str) -> str:
        """Clean and normalize text content"""
        # Remove excessive whitespace while preserving structure
        text = re.sub(r'\n\s*\n\s*\n+', '\n\n', text)  # Max 2 consecutive newlines
        text = re.sub(r'[ \t]+', ' ', text)  # Normalize spaces and tabs
        text = text.strip()
        return text

    def split_into_chunks(self, text: str, chunk_size: Optional[int] = None,
                         overlap: Optional[int] = None) -> List[str]:
        """
        Split text into chunks with overlap

        Args:
            text: The text to split
            chunk_size: Size of each chunk in tokens (defaults to setting)
            overlap: Overlap between chunks in tokens (defaults to setting)

        Returns:
            List of text chunks
        """
        if chunk_size is None:
            chunk_size = self.settings_manager.setting_get("rag_chunk_size") or 800
        if overlap is None:
            overlap = self.settings_manager.setting_get("rag_chunk_overlap") or 100

        if not text.strip():
            return []

        # Tokenize the entire text
        tokens = self.encoding.encode(text)

        if len(tokens) <= chunk_size:
            # Text is smaller than chunk size, return as single chunk
            return [text]

        chunks = []
        start = 0

        while start < len(tokens):
            # Calculate end position
            end = min(start + chunk_size, len(tokens))

            # Extract tokens for this chunk
            chunk_tokens = tokens[start:end]

            # Decode back to text
            chunk_text = self.encoding.decode(chunk_tokens)

            # Clean up chunk boundaries - try to end at sentence boundaries
            if end < len(tokens):  # Not the last chunk
                chunk_text = self._clean_chunk_boundary(chunk_text)

            chunks.append(chunk_text.strip())

            # Move start position (accounting for overlap)
            if end >= len(tokens):
                break  # We've reached the end

            start = end - overlap
            if start < 0:
                start = 0

        return [chunk for chunk in chunks if chunk.strip()]

    def _clean_chunk_boundary(self, chunk_text: str) -> str:
        """Try to end chunks at natural boundaries (sentences, paragraphs)"""
        # Try to end at paragraph boundary
        if '\n\n' in chunk_text:
            last_paragraph = chunk_text.rfind('\n\n')
            if last_paragraph > len(chunk_text) * 0.7:  # Only if we don't lose too much
                return chunk_text[:last_paragraph + 2]

        # Try to end at sentence boundary
        sentence_endings = ['. ', '! ', '? ', '.\n', '!\n', '?\n']
        best_end = -1

        for ending in sentence_endings:
            pos = chunk_text.rfind(ending)
            if pos > best_end and pos > len(chunk_text) * 0.7:  # Only if we don't lose too much
                best_end = pos + len(ending)

        if best_end > -1:
            return chunk_text[:best_end]

        # If no good boundary found, return as-is
        return chunk_text

    def estimate_line_numbers(self, full_text: str, chunk_text: str) -> Tuple[int, int]:
        """Estimate line numbers for a chunk within the full text"""
        try:
            # Find the chunk in the full text
            chunk_start = full_text.find(chunk_text)
            if chunk_start == -1:
                return 1, 1  # Fallback if chunk not found exactly

            # Count newlines before the chunk
            lines_before = full_text[:chunk_start].count('\n')
            lines_in_chunk = chunk_text.count('\n')

            start_line = lines_before + 1
            end_line = start_line + lines_in_chunk

            return start_line, end_line
        except:
            return 1, 1  # Fallback on any error

    def process_file(self, file_path: str, collection_name: str, collection_path: str = None) -> List[Dict[str, Any]]:
        """
        Process a single file into chunks with metadata

        Args:
            file_path: Path to the file
            collection_name: Name of the collection this file belongs to
            collection_path: Path to the collection root (for relative path calculation)

        Returns:
            List of chunk dictionaries with metadata
        """
        # Extract filename (use relative path if collection_path provided)
        if collection_path:
            filename = os.path.relpath(file_path, collection_path)
        else:
            filename = os.path.basename(file_path)

        # Load and clean content
        content = self.load_file(file_path)
        if not content:
            return []

        content = self.clean_text(content)
        if not content:
            return []

        # Split into chunks
        chunks = self.split_into_chunks(content)
        if not chunks:
            return []

        # Create chunk objects with metadata
        processed_chunks = []
        for i, chunk_text in enumerate(chunks):
            start_line, end_line = self.estimate_line_numbers(content, chunk_text)

            # Calculate file hash for tracking
            file_hash = self._get_file_hash(file_path)

            chunk_data = {
                "content": chunk_text,
                "filename": filename,
                "file_path": file_path,
                "source_file_path": file_path,  # For smart rebuild tracking
                "source_file_hash": file_hash,  # For smart rebuild tracking
                "collection_name": collection_name,
                "chunk_index": i,
                "total_chunks": len(chunks),
                "token_count": self.count_tokens(chunk_text),
                "start_line": start_line,
                "end_line": end_line,
                "created_at": None,  # Will be set when embeddings are generated
            }
            processed_chunks.append(chunk_data)

        return processed_chunks

    def _get_file_hash(self, file_path: str) -> str:
        """Calculate SHA-256 hash of a file for tracking changes"""
        try:
            hash_sha256 = hashlib.sha256()
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_sha256.update(chunk)
            return hash_sha256.hexdigest()
        except Exception as e:
            print_md(f"Error calculating hash for {file_path}: {e}")
            return ""

    def process_collection(self, collection_path: str) -> List[Dict[str, Any]]:
        """
        Process all files in a collection directory

        Args:
            collection_path: Path to the collection directory

        Returns:
            List of all chunks from all files in the collection
        """
        if not os.path.exists(collection_path):
            raise FileNotFoundError(f"Collection path does not exist: {collection_path}")

        collection_name = os.path.basename(collection_path)
        all_chunks = []

        # Process all supported files recursively in the directory
        for root, dirs, files in os.walk(collection_path):
            for filename in files:
                file_path = os.path.join(root, filename)

                if not is_supported_file(file_path):
                    continue

                # Get relative path for better display
                relative_path = os.path.relpath(file_path, collection_path)
                print_md(f"Processing {relative_path}...")

                try:
                    file_chunks = self.process_file(file_path, collection_name, collection_path)
                    all_chunks.extend(file_chunks)
                except Exception as e:
                    print_md(f"Error processing {relative_path}: {e}")
                    continue

        print_md(f"Processed {len(all_chunks)} chunks from {collection_name}")
        return all_chunks

    def get_collection_stats(self, collection_path: str) -> Dict[str, Any]:
        """Get statistics about a collection"""
        if not os.path.exists(collection_path):
            return {"error": "Collection not found"}

        stats = {
            "name": os.path.basename(collection_path),
            "path": collection_path,
            "file_count": 0,
            "supported_files": [],
            "unsupported_files": [],
            "total_size_bytes": 0
        }

        for root, dirs, files in os.walk(collection_path):
            for filename in files:
                file_path = os.path.join(root, filename)

                file_ext = os.path.splitext(filename)[1].lower()
                file_size = os.path.getsize(file_path)

                # Get relative path for better display
                relative_path = os.path.relpath(file_path, collection_path)

                if is_supported_file(file_path):
                    stats["supported_files"].append({
                        "name": relative_path,
                        "size_bytes": file_size,
                        "extension": file_ext
                    })
                    stats["total_size_bytes"] += file_size
                else:
                    stats["unsupported_files"].append({
                        "name": relative_path,
                        "extension": file_ext
                    })

        stats["file_count"] = len(stats["supported_files"])
        return stats